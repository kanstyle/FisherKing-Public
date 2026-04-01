#!/usr/bin/env python3
"""
Fire Emblem 8 Unit Stat Calculator
Parses event files and calculates unit stats based on character, class, and equipment data.
"""

import os
import re
import csv
import math
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class DefinitionResolver:
	"""Resolves definitions from TableEntryDefinitions.event and FE8Definitions.txt"""
	
	def __init__(self):
		self.definitions: Dict[str, str] = {}
	
	def load_definitions(self, filepath: str):
		"""Load definitions from a file"""
		with open(filepath, 'r', encoding='utf-8') as f:
			for line in f:
				line = line.strip()
				# Match #define lines
				match = re.match(r'#define\s+(\w+)\s+(.+)', line)
				if match:
					name, value = match.groups()
					self.definitions[name] = value.strip()
	
	def resolve(self, value: str) -> str:
		"""Resolve a definition to its final value (byte or hex)"""
		if not value:
			return value
		
		# If it's already a hex value, return it
		if value.startswith('0x') or value.startswith('0X'):
			return value
		
		# If it's a number, return it
		if value.isdigit():
			return value
		
		# Otherwise, try to resolve it
		seen = set()
		current = value
		
		while current in self.definitions and current not in seen:
			seen.add(current)
			next_val = self.definitions[current]
			
			# If we found a hex value or number, return it
			if next_val.startswith('0x') or next_val.startswith('0X'):
				return next_val
			if next_val.isdigit():
				return next_val
			
			current = next_val
		
		# Return the final value even if not fully resolved
		return current if current in self.definitions else value
	
	def to_int(self, value: str) -> int:
		"""Convert a resolved value to an integer"""
		resolved = self.resolve(value)
		if resolved.startswith('0x') or resolved.startswith('0X'):
			return int(resolved, 16)
		return int(resolved)


class UnitParser:
	"""Parses UNIT entries from event files"""
	
	UNIT_PATTERN = re.compile(
		r'UNIT\s+(\w+)\s+(\w+)\s+(\w+)\s+Level\((\d+),(\w+),(\w+)\)\s+'
		r'\[[\d,]+\]\s+\w+\s+\w+\s+\w+\s+\w+\s+'
		r'\[([^\]]+)\]'
	)
	
	@staticmethod
	def parse_unit(line: str) -> Optional[Dict]:
		"""Parse a UNIT line and extract relevant fields"""
		match = UnitParser.UNIT_PATTERN.search(line)
		if not match:
			return None
		
		char_id, class_id, _, level, allegiance, autolevel, items = match.groups()
		
		# Parse items
		item_list = [item.strip() for item in items.split(',')]
		# Filter out 0x0 entries
		item_list = [item for item in item_list if item != '0x0']
		
		return {
			'char_id': char_id,
			'class_id': class_id,
			'level': int(level),
			'allegiance': allegiance,
			'autolevel': autolevel.lower() == 'true',
			'items': item_list[:4]  # Max 4 items
		}
	
	@staticmethod
	def find_units_in_file(filepath: str) -> List[Dict]:
		"""Find all UNIT entries in a file"""
		units = []
		with open(filepath, 'r', encoding='utf-8') as f:
			for line in f:
				unit = UnitParser.parse_unit(line)
				if unit:
					units.append(unit)
		return units
	
	@staticmethod
	def parse_ch0_unit_groups(filepath: str) -> List[Tuple[str, List[Dict]]]:
		"""Parse unit groups from Ch0.event, preserving order.
		
		Returns:
			List of tuples: [(chapter_suffix, [unit_dicts]), ...]
			e.g., [("1", [...]), ("2", [...]), ("5a", [...]), ("Final", [...])]
		"""
		groups = []
		current_suffix = None
		current_units = []
		
		with open(filepath, 'r', encoding='utf-8') as f:
			for line in f:
				# Check for new chapter label: UnitsCh<suffix>:
				match = re.match(r'UnitsCh([^:]+):', line.strip())
				if match:
					# Save previous group if exists
					if current_suffix is not None:
						groups.append((current_suffix, current_units))
					# Start new group
					current_suffix = match.group(1)
					current_units = []
					continue
				
				# Check for blank UNIT terminator
				if line.strip() == 'UNIT':
					if current_suffix is not None:
						groups.append((current_suffix, current_units))
						current_suffix = None
						current_units = []
					continue
				
				# If we're in a group, try to parse unit
				if current_suffix is not None:
					unit = UnitParser.parse_unit(line)
					if unit is not None:
						current_units.append(unit)
		
		return groups
	
	@staticmethod
	def parse_autolevel_commands(filepath: str) -> Dict[str, Dict[str, int]]:
		"""Parse AutoLevel commands from Ch0.event
		
		Returns:
			{chapter_suffix: {char_id: target_level}}
			e.g., {'4': {'HeroChar': 6, 'Dragon': 6}, '5': {'HeroChar': 7, ...}}
		"""
		autolevel_map = {}
		current_chapter = None
		in_chapter_section = False
		
		with open(filepath, 'r', encoding='utf-8') as f:
			for line in f:
				line_stripped = line.strip()
				
				# Check for start of chapter section: SVAL 7 Ch<suffix>
				match = re.match(r'SVAL\s+7\s+Ch(.+)', line_stripped)
				if match:
					current_chapter = match.group(1)
					in_chapter_section = True
					autolevel_map[current_chapter] = {}
					continue
				
				# Check for end of chapter section: MNC2 Ch<suffix>
				if in_chapter_section:
					match = re.match(r'MNC2\s+Ch(.+)', line_stripped)
					if match and match.group(1) == current_chapter:
						in_chapter_section = False
						current_chapter = None
						continue
				
				# If we're in a chapter section, look for AutoLevel commands
				if in_chapter_section and current_chapter:
					# Match AutoLevel(CharID, Level)
					match = re.match(r'AutoLevel\((\w+),\s*(\d+)\)', line_stripped)
					if match:
						char_id = match.group(1)
						target_level = int(match.group(2))
						autolevel_map[current_chapter][char_id] = target_level
		
		return autolevel_map


def build_cumulative_rosters(groups_ordered: List[Tuple[str, List[Dict]]]) -> Dict[str, List[Dict]]:
	"""Build cumulative rosters based on order in Ch0.event.
	
	Args:
		groups_ordered: [(suffix, [units]), ...] in order from Ch0.event
	
	Returns:
		{suffix: [cumulative units up to and including this group], ...}
	"""
	cumulative = {}
	all_units = []
	
	for suffix, units in groups_ordered:
		all_units.extend(units)
		cumulative[suffix] = list(all_units)  # Copy
	
	return cumulative


def extract_chapter_suffix(filename: str) -> Optional[str]:
	"""Extract the chapter suffix from a filename.
	
	Args:
		filename: e.g., "Ch5.event", "Ch5a.event", "ChFinal.event"
	
	Returns:
		The suffix after "Ch" (without .event), or None if not a chapter file
		e.g., "5", "5a", "Final"
	"""
	# Remove .event extension
	name = filename[:-6] if filename.endswith('.event') else filename
	
	# Match Ch<anything>
	match = re.match(r'Ch(.+)', name)
	if match:
		return match.group(1)
	return None


class TableLoader:
	"""Loads and queries CSV tables"""
	
	def __init__(self):
		self.tables: Dict[str, List[Dict]] = {}
	
	def load_csv(self, filepath: str, table_name: str):
		"""Load a CSV file into memory"""
		rows = []
		with open(filepath, 'r', encoding='utf-8') as f:
			reader = csv.DictReader(f)
			for row in reader:
				rows.append(row)
		self.tables[table_name] = rows
	
	def find_by_id(self, table_name: str, id_value: str, id_column: str = 'ID') -> Optional[Dict]:
		"""Find a row in a table by ID"""
		if table_name not in self.tables:
			return None
		
		# Normalize hex values for comparison
		def normalize_id(val):
			if not val:
				return val
			val = val.strip()
			if val.startswith('0x') or val.startswith('0X'):
				try:
					# Convert to int and back to normalized hex
					return hex(int(val, 16))
				except:
					return val
			return val
		
		normalized_search = normalize_id(id_value)
		
		for row in self.tables[table_name]:
			row_id = normalize_id(row.get(id_column, ''))
			if row_id == normalized_search:
				return row
		return None
	
	def find_by_row_number(self, table_name: str, row_number: int) -> Optional[Dict]:
		"""Find a row in a table by row number (0-indexed, after header)"""
		if table_name not in self.tables:
			return None
		if row_number < 0 or row_number >= len(self.tables[table_name]):
			return None
		return self.tables[table_name][row_number]
	
	def find_character_by_hex_id(self, hex_id: str) -> Optional[Dict]:
		"""Find a character by hex ID in the first column (0x803d64)"""
		if 'character' not in self.tables:
			return None
		
		# Normalize the hex ID
		normalized_search = hex_id.strip().lower()
		if not normalized_search.startswith('0x'):
			return None
		
		for row in self.tables['character']:
			first_col = row.get('0x803d64', '').strip()
			if first_col:
				# Extract hex ID from format "0xNN Name"
				parts = first_col.split()
				if parts and parts[0].lower() == normalized_search:
					return row
		
		return None
	
	def get_stat_value(self, row: Dict, column: str, default: int = 0) -> int:
		"""Get a stat value from a row, handling empty strings"""
		value = row.get(column, '').strip()
		if not value or value == '':
			return default
		try:
			return int(value)
		except ValueError:
			return default


class StatCalculator:
	"""Calculates unit stats based on character, class, and equipment"""
	
	def __init__(self, resolver: DefinitionResolver, table_loader: TableLoader):
		self.resolver = resolver
		self.tables = table_loader
	
	@staticmethod
	def parse_difficulty_bonus(difficulty_hex: str) -> dict:
		"""
		Parse difficulty bonus from hex format: 0x[Normal][Hard][Easy]
		
		Format interpretation:
		- Position 0 (Normal): Autolevels to SUBTRACT (e.g., 2 = -2 autolevels)
		- Position 1 (Hard): Autolevels to ADD (e.g., 3 = +3 autolevels)
		- Position 2 (Easy): Autolevels to SUBTRACT (e.g., 4 = -4 autolevels)
		
		Each digit can be 0-F (0-15).
		
		Example: 0x234 = Normal: -2, Hard: +3, Easy: -4
		Example: 0x050 = Normal: 0, Hard: +5, Easy: 0
		"""
		if not difficulty_hex or not difficulty_hex.startswith('0x'):
			return {'Normal': 0, 'Hard': 0, 'Easy': 0}
		
		try:
			# Remove 0x prefix and ensure we have 3 digits
			hex_str = difficulty_hex[2:].strip().upper()
			if len(hex_str) < 3:
				hex_str = hex_str.zfill(3)
			
			# Parse each digit
			normal_hex = hex_str[0]
			hard_hex = hex_str[1]
			easy_hex = hex_str[2]
			
			# Normal and Easy are subtractions, Hard is addition
			normal_val = -int(normal_hex, 16)  # Negative (subtract)
			hard_val = int(hard_hex, 16)       # Positive (add)
			easy_val = -int(easy_hex, 16)      # Negative (subtract)
			
			return {
				'Normal': normal_val,
				'Hard': hard_val,
				'Easy': easy_val
			}
		except (ValueError, IndexError):
			return {'Normal': 0, 'Hard': 0, 'Easy': 0}
	
	def calculate_unit_stats(self, unit: Dict, chapter_name: str, difficulty_bonus: int = 0, 
	                         autolevel_map: Dict[str, Dict[str, int]] = None, 
	                         chapter_suffix: str = None) -> Dict:
		"""Calculate all stats for a unit
		
		Args:
			unit: Unit data dictionary
			chapter_name: Name of the chapter
			difficulty_bonus: Level adjustment from difficulty mode (added to chapLevel for calculations)
			autolevel_map: Dictionary of AutoLevel commands from Ch0.event
			chapter_suffix: Chapter suffix (e.g., "4", "5a") for looking up AutoLevel commands
		"""
		# For character lookup, try the string name first, then resolved hex ID
		char_name = unit['char_id']
		char_id_hex = self.resolver.resolve(unit['char_id'])
		
		# For class lookup, try string name first, then resolved hex ID
		class_name = unit['class_id']
		class_id_hex = self.resolver.resolve(unit['class_id'])
		
		# Get character and class data
		# For characters: try name first, then hex ID in ID column, then hex ID in first column
		char_data = self.tables.find_by_id('character', char_name)
		if not char_data:
			char_data = self.tables.find_by_id('character', char_id_hex)
		if not char_data:
			# Try searching by hex ID in the first column (for entries like "0x80 Soldier")
			char_data = self.tables.find_character_by_hex_id(char_id_hex)
		
		class_data = self.tables.find_by_id('class', class_name)
		if not class_data:
			# Try hex ID
			class_data = self.tables.find_by_id('class', class_id_hex)
		
		if not char_data or not class_data:
			return None
		
		# For mag tables, extract the hex ID from the first column of the main table
		# The first column format is "0xNN Name" - extract the hex part
		char_first_col = char_data.get('0x803d64', '').strip()
		char_hex_id = None
		if char_first_col:
			# Extract hex ID (format: "0x01 Hero" -> "0x01")
			parts = char_first_col.split()
			if parts and parts[0].startswith('0x'):
				try:
					char_hex_id = int(parts[0], 16)
				except:
					pass
		
		class_first_col = class_data.get('INLINE NewClassTable', '').strip()
		class_hex_id = None
		if class_first_col:
			# For classes, some have string IDs, some have hex
			# Try to parse as hex if it starts with 0x
			if class_first_col.startswith('0x'):
				try:
					class_hex_id = int(class_first_col.split()[0], 16)
				except:
					pass
			else:
				# String ID like "Veteran" - need to resolve it
				try:
					resolved = self.resolver.resolve(class_name)
					if resolved and resolved.startswith('0x'):
						class_hex_id = int(resolved, 16)
				except:
					pass
		
		# Use hex ID as row index into mag tables
		mag_char_data = None
		if char_hex_id is not None:
			mag_char_data = self.tables.find_by_row_number('mag_char', char_hex_id)
		
		mag_class_data = None
		if class_hex_id is not None:
			mag_class_data = self.tables.find_by_row_number('mag_class', class_hex_id)
		
		# Get base stats
		char_bases = {
			'HP': self.tables.get_stat_value(char_data, 'HP'),
			'Atk': self.tables.get_stat_value(char_data, 'Atk'),
			'Skl': self.tables.get_stat_value(char_data, 'Skl'),
			'Spd': self.tables.get_stat_value(char_data, 'Spd'),
			'Def': self.tables.get_stat_value(char_data, 'Def'),
			'Res': self.tables.get_stat_value(char_data, 'Res'),
			'Luck': self.tables.get_stat_value(char_data, 'Luck'),
			'Con': self.tables.get_stat_value(char_data, 'Con'),
		}
		
		class_bases = {
			'HP': self.tables.get_stat_value(class_data, 'HP'),
			'Atk': self.tables.get_stat_value(class_data, 'Atk'),
			'Skl': self.tables.get_stat_value(class_data, 'Skl'),
			'Spd': self.tables.get_stat_value(class_data, 'Spd'),
			'Def': self.tables.get_stat_value(class_data, 'Def'),
			'Res': self.tables.get_stat_value(class_data, 'Res'),
			'Con': self.tables.get_stat_value(class_data, 'Con'),
			'Luck': 0,  # Classes don't have Luck bases
		}
		
		# Get magic bases
		char_mag = self.tables.get_stat_value(mag_char_data, 'Base Magic') if mag_char_data else 0
		class_mag = self.tables.get_stat_value(mag_class_data, 'Base Magic') if mag_class_data else 0
		
		# Get growths
		class_growths = {
			'HP': self.tables.get_stat_value(class_data, 'HP Growth'),
			'Atk': self.tables.get_stat_value(class_data, 'Atk Growth'),
			'Skl': self.tables.get_stat_value(class_data, 'Skl Growth'),
			'Spd': self.tables.get_stat_value(class_data, 'Spd Growth'),
			'Def': self.tables.get_stat_value(class_data, 'Def Growth'),
			'Res': self.tables.get_stat_value(class_data, 'Res Growth'),
			'Luck': self.tables.get_stat_value(class_data, 'Luck Growth'),
		}
		
		class_mag_growth = self.tables.get_stat_value(mag_class_data, 'Magic Growth') if mag_class_data else 0
		
		# Check if promoted
		class_ability = class_data.get('CharClassAbility', '')
		is_promoted = 'IsPromoted' in class_ability
		
		# Calculate final stats
		level = unit['level']
		allegiance = unit['allegiance']
		autolevel = unit['autolevel']
		
		# Apply difficulty bonus to level for calculations (but keep original level for output)
		adjusted_level = level + difficulty_bonus
		
		stats = {}
		
		if allegiance in ['Enemy', 'NPC']:
			# Enemy/NPC units use only class growths
			for stat in ['HP', 'Atk', 'Skl', 'Spd', 'Def', 'Res', 'Luck']:
				base = class_bases[stat] + char_bases[stat]
				if autolevel:
					growth_bonus = class_growths[stat] * adjusted_level // 100
					if is_promoted:
						growth_bonus += class_growths[stat] * 10 // 100
					base += growth_bonus
				stats[stat] = base
			
			# Magic
			mag_base = class_mag + char_mag
			if autolevel:
				mag_growth_bonus = class_mag_growth * adjusted_level // 100
				if is_promoted:
					mag_growth_bonus += class_mag_growth * 10 // 100
				mag_base += mag_growth_bonus
			stats['Mag'] = mag_base
			
		else:  # Ally
			# Ally units use character growths
			char_growths = {
				'HP': self.tables.get_stat_value(char_data, 'HP Growth'),
				'Atk': self.tables.get_stat_value(char_data, 'Atk Growth'),
				'Skl': self.tables.get_stat_value(char_data, 'Skl Growth'),
				'Spd': self.tables.get_stat_value(char_data, 'Spd Growth'),
				'Def': self.tables.get_stat_value(char_data, 'Def Growth'),
				'Res': self.tables.get_stat_value(char_data, 'Res Growth'),
				'Luck': self.tables.get_stat_value(char_data, 'Luck Growth'),
			}
			
			char_mag_growth = self.tables.get_stat_value(mag_char_data, 'Magic Growth') if mag_char_data else 0
			
			# Check if this character has an AutoLevel command for this chapter
			should_autolevel = False
			levels_to_gain = 0
			
			if autolevel_map and chapter_suffix:
				chapter_autolevels = autolevel_map.get(chapter_suffix, {})
				target_level = None
				
				# Check both char_name and char_id_hex
				if char_name in chapter_autolevels:
					target_level = chapter_autolevels[char_name]
				# Also check the ID column value
				elif char_data.get('ID', '') in chapter_autolevels:
					target_level = chapter_autolevels[char_data.get('ID', '')]
				
				if target_level is not None:
					# Calculate levels to gain: from base level to target level
					levels_to_gain = target_level - level
					if levels_to_gain > 0:
						should_autolevel = True
			
			for stat in ['HP', 'Atk', 'Skl', 'Spd', 'Def', 'Res', 'Luck']:
				base = class_bases[stat] + char_bases[stat]
				if should_autolevel:
					growth_bonus = char_growths[stat] * levels_to_gain // 100
					base += growth_bonus
				stats[stat] = base
			
			# Magic
			mag_base = class_mag + char_mag
			if should_autolevel:
				mag_growth_bonus = char_mag_growth * levels_to_gain // 100
				mag_base += mag_growth_bonus
			stats['Mag'] = mag_base
		
		# Con is always base only
		stats['Con'] = class_bases['Con'] + char_bases['Con']
		stats['Lv'] = level
		
		# Prepare result
		result = {
			'UnitID': f"{unit['char_id']}_{unit['class_id']}",
			'Chapter': chapter_name,
			'Allegiance': allegiance,  # Add allegiance for color coding
			'Lv': stats['Lv'],
			'HP': stats['HP'],
			'Str': stats['Atk'],  # 'Atk' in game tables = 'Str' in output format
			'Mag': stats['Mag'],
			'Skl': stats['Skl'],
			'Spd': stats['Spd'],
			'Def': stats['Def'],
			'Res': stats['Res'],
			'Luck': stats['Luck'],
			'Con': stats['Con'],
			'weapons': []
		}
		
		# Calculate weapon stats
		for item in unit['items']:
			weapon_stats = self.calculate_weapon_stats(item, result)
			if weapon_stats:
				result['weapons'].append(weapon_stats)
		
		return result
	
	def calculate_weapon_stats(self, item_id: str, unit_result: Dict) -> Optional[Dict]:
		"""Calculate stats for a weapon"""
		# Resolve item ID
		resolved_id = self.resolver.resolve(item_id)
		
		# Get item data
		item_data = self.tables.find_by_id('item', resolved_id)
		if not item_data:
			return None
		
		# Check if it's a weapon
		item_ability = item_data.get('ItemAbility', '')
		if 'IsWeapon' not in item_ability:
			return None
		
		# Get weapon stats
		mt = self.tables.get_stat_value(item_data, 'Mt')
		hit = self.tables.get_stat_value(item_data, 'Hit')
		wt = self.tables.get_stat_value(item_data, 'Wt')
		crit = self.tables.get_stat_value(item_data, 'Crit')
		
		# Calculate derived stats
		is_magic = 'MagicDamage' in item_ability or 'IsMagic' in item_ability
		atk = mt + (unit_result['Mag'] if is_magic else unit_result['Str'])

		weapon_hit = 2 * unit_result['Skl'] + unit_result['Luck'] + hit
		weapon_crit = crit

		# AS = Spd - max(0, Wt - Con)
		wt_penalty = max(0, wt - unit_result['Con'])
		attack_speed = unit_result['Spd'] - wt_penalty

		avoid = 2 * unit_result['Spd'] + unit_result['Luck']

		wexp = self.tables.get_stat_value(item_data, 'Weapon EXP')
		item_type = item_data.get('Type', '').strip()
		is_bomb = (item_type == 'Special')

		return {
			'Weapon': item_id,
			'Atk': atk,
			'Hit': weapon_hit,
			'Crit': weapon_crit,
			'AS': attack_speed,
			'Avoid': avoid,
			'WExp': wexp,
			'IsMagic': is_magic,
			'IsBomb': is_bomb,
		}


def simulate_chapter_wexp(ally_units: List[Dict], enemy_units: List[Dict]) -> Dict[str, Dict[str, int]]:
	"""
	Simulate turn-by-turn combat and calculate weapon EXP gained by each ally.

	Each turn every ally is assigned the enemy they can kill in the fewest total attacks.
	When multiple allies want the same enemy, each gets a distinct target (greedy first-come
	assignment in roster order).  If no weapon can damage an enemy the ally is skipped.

	Doubling rule: ally attacks twice per engagement when ally_weapon_AS >= enemy_AS + 4.
	Bomb rule: if any non-bomb weapon can kill in one engagement, bombs are excluded from
	           weapon selection for that matchup.

	Returns:
		{ally_unit_id: {weapon_item_id: total_wexp_gained}}
	"""
	if not ally_units or not enemy_units:
		return {}

	# Build enemy instance pool expanded by Count
	enemy_pool = []
	for enemy in enemy_units:
		count = enemy.get('Count', 1)
		enemy_as = enemy['weapons'][0]['AS'] if enemy.get('weapons') else enemy.get('Spd', 0)
		for _ in range(count):
			enemy_pool.append({
				'hp':  enemy['HP'],
				'def': enemy['Def'],
				'res': enemy['Res'],
				'as':  enemy_as,
			})

	# WEXP accumulator: {ally_unit_id: {weapon_item_id: total}}
	wexp_tracker: Dict[str, Dict[str, int]] = {ally['UnitID']: {} for ally in ally_units}

	def best_weapon_vs(ally: Dict, enemy_inst: Dict):
		"""
		Choose the best weapon for ally to use against enemy_inst.

		Returns (weapon_dict, attacks_per_engagement, total_attacks_to_kill)
		or None if no weapon can deal damage.
		"""
		candidates = []
		for weapon in ally.get('weapons', []):
			if not weapon.get('WExp', 0):
				continue  # non-weapon item
			dmg = max(0, weapon['Atk'] - (enemy_inst['res'] if weapon.get('IsMagic') else enemy_inst['def']))
			if dmg <= 0:
				continue
			ape = 2 if weapon['AS'] >= enemy_inst['as'] + 4 else 1  # attacks per engagement
			total_atk = math.ceil(enemy_inst['hp'] / dmg)            # total hits to kill
			eng_to_kill = math.ceil(total_atk / ape)                 # engagements to kill
			candidates.append({
				'weapon':      weapon,
				'ape':         ape,
				'eng_to_kill': eng_to_kill,
				'is_bomb':     weapon.get('IsBomb', False),
			})

		if not candidates:
			return None

		# Bomb rule: if any non-bomb can kill in 1 engagement, exclude all bombs
		if any(not c['is_bomb'] and c['eng_to_kill'] == 1 for c in candidates):
			candidates = [c for c in candidates if not c['is_bomb']]

		# Pick minimum engagements; prefer non-bomb on ties
		candidates.sort(key=lambda c: (c['eng_to_kill'], c['is_bomb']))
		best = candidates[0]
		return best['weapon'], best['ape'], best['eng_to_kill']

	# Safety cap: no chapter should ever need more turns than 10× total enemies
	max_turns = sum(e.get('Count', 1) for e in enemy_units) * 10
	turn = 0

	while enemy_pool and turn < max_turns:
		turn += 1
		assigned: set = set()

		for ally in ally_units:
			if not enemy_pool:
				break

			# Find unassigned enemy with the fewest engagements to kill
			best_idx = -1
			best_eng  = float('inf')
			best_wpn  = None
			best_ape  = 1

			for idx, enemy_inst in enumerate(enemy_pool):
				if idx in assigned:
					continue
				result = best_weapon_vs(ally, enemy_inst)
				if result is None:
					continue
				weapon, ape, eng = result
				if eng < best_eng:
					best_eng = eng
					best_idx = idx
					best_wpn  = weapon
					best_ape  = ape

			if best_idx < 0:
				continue  # ally cannot damage any remaining enemy

			assigned.add(best_idx)
			enemy_inst = enemy_pool[best_idx]

			# Resolve actual damage and HP change
			dmg = max(0, best_wpn['Atk'] - (enemy_inst['res'] if best_wpn.get('IsMagic') else enemy_inst['def']))
			enemy_inst['hp'] -= dmg * best_ape

			# Accumulate WEXP
			gained = best_wpn['WExp'] * best_ape
			wpn_id  = best_wpn['Weapon']
			ally_id = ally['UnitID']
			wexp_tracker[ally_id][wpn_id] = wexp_tracker[ally_id].get(wpn_id, 0) + gained

		# Remove enemies that have been killed
		enemy_pool = [e for e in enemy_pool if e['hp'] > 0]

	return wexp_tracker


def export_to_excel(results: Dict[str, List[Dict]], output_path: str):
	"""Export calculated stats to Excel with one sheet per chapter
	
	Features:
	- Units separated by allegiance (Ally, NPC, Enemy)
	- Color-coded rows (light blue for Ally, light green for NPC, light red for Enemy)
	- Additional columns: AS+4, Mag OHKO, Phys OHKO, Mag ORKO, Phys ORKO
	"""
	wb = openpyxl.Workbook()
	wb.remove(wb.active)  # Remove default sheet
	
	# Header style
	header_font = Font(bold=True)
	header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
	
	# Allegiance colors
	ally_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
	npc_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # Light green
	enemy_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid") # Light red
	
	for chapter_name in sorted(results.keys()):
		units = results[chapter_name]
		if not units:
			continue
		
		# Separate units by allegiance
		ally_units = [u for u in units if u.get('Allegiance') == 'Ally']
		npc_units = [u for u in units if u.get('Allegiance') == 'NPC']
		enemy_units = [u for u in units if u.get('Allegiance') == 'Enemy']
		
		# Create sheet
		ws = wb.create_sheet(title=chapter_name[:31])  # Excel sheet name limit
		
		# Write header
		# Col:  1       2    3     4      5      6      7      8      9      10      11     12
		#       UnitID  Lv   HP    Str    Mag    Skl    Spd    Def    Res    Luck    Con    Count
		# Col:  13      14   15    16     17     18     19     20     21          22          23          24         25
		#       Weapon  Atk  Hit   Crit   AS     Avoid  AS+4   WEXP   Mag OHKO    Phys OHKO   Bomb OHKO   Mag ORKO   Phys ORKO
		headers = ['UnitID', 'Lv', 'HP', 'Str', 'Mag', 'Skl', 'Spd', 'Def', 'Res', 'Luck', 'Con', 'Count',
				   'Weapon', 'Atk', 'Hit', 'Crit', 'AS', 'Avoid', 'AS+4', 'WEXP',
				   'Mag OHKO', 'Phys OHKO', 'Bomb OHKO', 'Mag ORKO', 'Phys ORKO']
		
		for col_idx, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col_idx, value=header)
			cell.font = header_font
			cell.fill = header_fill
		
		row_idx = 2
		
		# Helper function to write unit rows with color coding
		def write_unit_section(unit_list, fill_color):
			nonlocal row_idx
			for unit in unit_list:
				# Use actual HP (not HP Cap) for OHKO/ORKO calculations
				hp = unit['HP']
				def_stat = unit['Def']
				res_stat = unit['Res']
				
				if not unit['weapons']:
					# Unit with no weapons
					ws.cell(row=row_idx, column=1,  value=unit['UnitID'])
					ws.cell(row=row_idx, column=2,  value=unit['Lv'])
					ws.cell(row=row_idx, column=3,  value=hp)
					ws.cell(row=row_idx, column=4,  value=unit['Str'])
					ws.cell(row=row_idx, column=5,  value=unit['Mag'])
					ws.cell(row=row_idx, column=6,  value=unit['Skl'])
					ws.cell(row=row_idx, column=7,  value=unit['Spd'])
					ws.cell(row=row_idx, column=8,  value=def_stat)
					ws.cell(row=row_idx, column=9,  value=res_stat)
					ws.cell(row=row_idx, column=10, value=unit['Luck'])
					ws.cell(row=row_idx, column=11, value=unit['Con'])
					ws.cell(row=row_idx, column=12, value=unit.get('Count'))  # Count (enemies only)
					# Cols 13-20 are weapon stats / WEXP (empty for weaponless units)
					# Cols 21-25: OHKO/ORKO stats
					ws.cell(row=row_idx, column=21, value=hp + res_stat)           # Mag OHKO
					ws.cell(row=row_idx, column=22, value=hp + def_stat)           # Phys OHKO
					ws.cell(row=row_idx, column=23, value=int(hp * 0.75) + res_stat)  # Bomb OHKO
					ws.cell(row=row_idx, column=24, value=hp // 2 + res_stat)      # Mag ORKO
					ws.cell(row=row_idx, column=25, value=hp // 2 + def_stat)      # Phys ORKO

					# Apply color to entire row
					for col in range(1, 26):
						ws.cell(row=row_idx, column=col).fill = fill_color

					row_idx += 1
				else:
					# Unit with weapons
					wexp_map = unit.get('wexp_by_weapon', {})  # {weapon_id: total_wexp} for allies
					for weapon in unit['weapons']:
						ws.cell(row=row_idx, column=1,  value=unit['UnitID'])
						ws.cell(row=row_idx, column=2,  value=unit['Lv'])
						ws.cell(row=row_idx, column=3,  value=hp)
						ws.cell(row=row_idx, column=4,  value=unit['Str'])
						ws.cell(row=row_idx, column=5,  value=unit['Mag'])
						ws.cell(row=row_idx, column=6,  value=unit['Skl'])
						ws.cell(row=row_idx, column=7,  value=unit['Spd'])
						ws.cell(row=row_idx, column=8,  value=def_stat)
						ws.cell(row=row_idx, column=9,  value=res_stat)
						ws.cell(row=row_idx, column=10, value=unit['Luck'])
						ws.cell(row=row_idx, column=11, value=unit['Con'])
						ws.cell(row=row_idx, column=12, value=unit.get('Count'))  # Count (enemies only)
						ws.cell(row=row_idx, column=13, value=weapon['Weapon'])
						ws.cell(row=row_idx, column=14, value=weapon['Atk'])
						ws.cell(row=row_idx, column=15, value=weapon['Hit'])
						ws.cell(row=row_idx, column=16, value=weapon['Crit'])
						ws.cell(row=row_idx, column=17, value=weapon['AS'])
						ws.cell(row=row_idx, column=18, value=weapon['Avoid'])
						ws.cell(row=row_idx, column=19, value=weapon['AS'] + 4)   # AS+4
						ws.cell(row=row_idx, column=20, value=wexp_map.get(weapon['Weapon']))  # WEXP (allies only)
						ws.cell(row=row_idx, column=21, value=hp + res_stat)           # Mag OHKO
						ws.cell(row=row_idx, column=22, value=hp + def_stat)           # Phys OHKO
						ws.cell(row=row_idx, column=23, value=int(hp * 0.75) + res_stat)  # Bomb OHKO
						ws.cell(row=row_idx, column=24, value=hp // 2 + res_stat)      # Mag ORKO
						ws.cell(row=row_idx, column=25, value=hp // 2 + def_stat)      # Phys ORKO

						# Apply color to entire row
						for col in range(1, 26):
							ws.cell(row=row_idx, column=col).fill = fill_color

						row_idx += 1
		
		# Write units by allegiance with spacing
		if ally_units:
			write_unit_section(ally_units, ally_fill)
			row_idx += 1  # Add blank row separator
		
		if npc_units:
			write_unit_section(npc_units, npc_fill)
			row_idx += 1  # Add blank row separator
		
		if enemy_units:
			write_unit_section(enemy_units, enemy_fill)
		
		# Auto-adjust column widths
		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = min(max_length + 2, 50)
			ws.column_dimensions[column_letter].width = adjusted_width
	
	wb.save(output_path)


def main():
	"""Main function"""
	import sys
	import os
	
	# Get base path from command line or use current directory
	if len(sys.argv) > 1:
		base_path = sys.argv[1]
	else:
		# Use current directory as base
		base_path = os.getcwd()
	
	# File paths with new defaults
	table_defs_path = os.path.join(base_path, 'Definitions', 'TableEntryDefinitions.event')
	fe8_defs_path = os.path.join(base_path, 'EventAssembler', 'EAStandardLibrary', 'FE8 Definitions.txt')
	chapter_events_dir = os.path.join(base_path, 'Events', 'ChapterEvents')
	char_class_dir = os.path.join(base_path, 'Tables', 'NightmareModules', 'CharactersClasses')
	items_dir = os.path.join(base_path, 'Tables', 'NightmareModules', 'Items')
	chapter_data_dir = os.path.join(base_path, 'Tables', 'NightmareModules', 'ChapterData')
	
	char_table_path = os.path.join(char_class_dir, 'CharacterTable.csv')
	class_table_path = os.path.join(char_class_dir, 'ClassTable.csv')
	mag_char_path = os.path.join(char_class_dir, 'MagCharEditor.csv')
	mag_class_path = os.path.join(char_class_dir, 'MagClassEditor.csv')
	item_table_path = os.path.join(items_dir, 'ItemTable.csv')
	chapter_data_path = os.path.join(chapter_data_dir, 'ChapterData.csv')
	
	# Check if files exist, if not try flat structure fallback
	if not os.path.exists(table_defs_path):
		print("Standard structure not found, trying flat structure...")
		table_defs_path = os.path.join(base_path, 'TableEntryDefinitions.event')
		fe8_defs_path = os.path.join(base_path, 'FE8 Definitions.txt')
		chapter_events_dir = base_path  # Look in base path
		char_table_path = os.path.join(base_path, 'CharacterTable.csv')
		class_table_path = os.path.join(base_path, 'ClassTable.csv')
		mag_char_path = os.path.join(base_path, 'MagCharEditor.csv')
		mag_class_path = os.path.join(base_path, 'MagClassEditor.csv')
		item_table_path = os.path.join(base_path, 'ItemTable.csv')
		chapter_data_path = os.path.join(base_path, 'ChapterData.csv')
	
	# Check if files exist
	required_files = [
		table_defs_path, fe8_defs_path, char_table_path, class_table_path,
		mag_char_path, mag_class_path, item_table_path
	]
	
	# ChapterData is optional
	has_chapter_data = os.path.exists(chapter_data_path)
	if not has_chapter_data:
		print("Warning: ChapterData.csv not found - difficulty modes will not be available")
	
	missing_files = [f for f in required_files if not os.path.exists(f)]
	if missing_files:
		print("ERROR: Required files not found:")
		for f in missing_files:
			print(f"  - {f}")
		print(f"\nBase path: {base_path}")
		print("\nExpected structure (standard):")
		print("  \\Definitions\\TableEntryDefinitions.event")
		print("  \\EventAssembler\\EAStandardLibrary\\FE8 Definitions.txt")
		print("  \\Events\\ChapterEvents\\")
		print("  \\Tables\\NightmareModules\\CharactersClasses\\")
		print("  \\Tables\\NightmareModules\\Items\\")
		print("  \\Tables\\NightmareModules\\ChapterData\\ (optional)")
		print("\nOr flat structure:")
		print("  \\TableEntryDefinitions.event")
		print("  \\FE8 Definitions.txt")
		print("  \\CharacterTable.csv, ClassTable.csv, etc.")
		print("  \\ChapterData.csv (optional)")
		return
	
	# Initialize components
	print("Loading definitions...")
	resolver = DefinitionResolver()
	resolver.load_definitions(table_defs_path)
	resolver.load_definitions(fe8_defs_path)
	
	print("Loading tables...")
	table_loader = TableLoader()
	table_loader.load_csv(char_table_path, 'character')
	table_loader.load_csv(class_table_path, 'class')
	table_loader.load_csv(mag_char_path, 'mag_char')
	table_loader.load_csv(mag_class_path, 'mag_class')
	table_loader.load_csv(item_table_path, 'item')
	
	# Load chapter data if available
	chapter_data_map = {}
	if has_chapter_data:
		table_loader.load_csv(chapter_data_path, 'chapter_data')
		# Build a map of EventsID -> chapter data
		for row in table_loader.tables['chapter_data']:
			events_id = row.get('EventsID', '').strip()
			if events_id:
				chapter_data_map[events_id] = row
	
	calculator = StatCalculator(resolver, table_loader)
	
	# Parse Ch0.event for ally unit groups
	print("Parsing Ch0.event for ally unit groups...")
	ch0_path = os.path.join(chapter_events_dir, 'Ch0.event')
	ally_rosters = {}
	autolevel_map = {}
	if os.path.exists(ch0_path):
		groups_ordered = UnitParser.parse_ch0_unit_groups(ch0_path)
		ally_rosters = build_cumulative_rosters(groups_ordered)
		autolevel_map = UnitParser.parse_autolevel_commands(ch0_path)
		print(f"  Found {len(groups_ordered)} unit groups in Ch0.event")
		for suffix, _ in groups_ordered:
			print(f"    - UnitsCh{suffix}")
		# Print AutoLevel commands found
		if autolevel_map:
			print(f"  Found AutoLevel commands for {len(autolevel_map)} chapters")
	else:
		print("Warning: Ch0.event not found - no ally units will be processed")
	
	# Parse units from all chapter event files
	print("Parsing chapter events...")
	results = {}
	
	# Determine which difficulty modes to process
	include_easy = False  # Default: off
	# Check for command line argument
	if len(sys.argv) > 2 and sys.argv[2].lower() in ['--easy', '-e', 'easy']:
		include_easy = True
	
	difficulty_modes = ['Normal', 'Hard']
	if include_easy:
		difficulty_modes.append('Easy')
	
	if os.path.isdir(chapter_events_dir):
		# Process all .event files in the directory
		event_files = [f for f in os.listdir(chapter_events_dir) if f.endswith('.event')]
		
		for filename in sorted(event_files):
			chapter_suffix = extract_chapter_suffix(filename)
			
			# Skip Ch0 and non-chapter files
			if chapter_suffix is None or chapter_suffix == '0':
				continue
			
			chapter_file = os.path.join(chapter_events_dir, filename)
			chapter_name = filename[:-6]  # For display/sheet naming
			
			# Get ally units from cumulative roster (if this chapter exists in Ch0)
			ally_units = ally_rosters.get(chapter_suffix, [])
			
			# Get enemy/NPC units from chapter file
			all_units = UnitParser.find_units_in_file(chapter_file)
			enemy_npc_units = [u for u in all_units if u['allegiance'] in ['Enemy', 'NPC']]

			# Count enemy instances before deduplication
			enemy_counts = {}
			for unit in enemy_npc_units:
				key = f"{unit['char_id']}_{unit['class_id']}"
				enemy_counts[key] = enemy_counts.get(key, 0) + 1

			# Combine
			units = ally_units + enemy_npc_units
			
			# Remove duplicates
			unique_units = {}
			for unit in units:
				key = f"{unit['char_id']}_{unit['class_id']}"
				if key not in unique_units:
					unique_units[key] = unit
			
			units = list(unique_units.values())
			
			if units:
				ally_count = len([u for u in units if u['allegiance'] == 'Ally'])
				enemy_npc_count = len(units) - ally_count
				print(f"  {chapter_name}: {ally_count} ally, {enemy_npc_count} enemy/NPC units")
				
				# Get chapter data for difficulty bonuses
				events_id = f"{chapter_name}EventsID"
				chapter_data = chapter_data_map.get(events_id, {})
				difficulty_hex = chapter_data.get('DifficultyBonus', '0x0')
				difficulty_bonuses = calculator.parse_difficulty_bonus(difficulty_hex)
				
				# Calculate stats for each difficulty mode
				for difficulty in difficulty_modes:
					sheet_name = f"{chapter_name} {difficulty}"
					difficulty_bonus = difficulty_bonuses.get(difficulty, 0)
					
					results[sheet_name] = []

					for unit in units:
						stats = calculator.calculate_unit_stats(unit, chapter_name, difficulty_bonus,
						                                        autolevel_map, chapter_suffix)
						if stats:
							if stats['Allegiance'] in ['Enemy', 'NPC']:
								unit_key = f"{unit['char_id']}_{unit['class_id']}"
								stats['Count'] = enemy_counts.get(unit_key, 1)
							results[sheet_name].append(stats)

					# Run WEXP simulation for this sheet
					ally_stats = [u for u in results[sheet_name] if u.get('Allegiance') == 'Ally']
					enemy_stats = [u for u in results[sheet_name] if u.get('Allegiance') in ['Enemy', 'NPC']]
					wexp_results = simulate_chapter_wexp(ally_stats, enemy_stats)
					for unit in results[sheet_name]:
						if unit.get('Allegiance') == 'Ally':
							unit['wexp_by_weapon'] = wexp_results.get(unit['UnitID'], {})

	if not results:
		print("ERROR: No chapter event files found or no units parsed!")
		return
	
	# Export to Excel in the same directory as the script
	script_dir = os.path.dirname(os.path.abspath(__file__))
	output_path = os.path.join(script_dir, 'unit_stats.xlsx')
	
	print(f"Calculating stats...")
	print(f"Exporting to {output_path}...")
	export_to_excel(results, output_path)
	print("Done!")
	print(f"\nOutput file: {output_path}")


if __name__ == '__main__':
	main()