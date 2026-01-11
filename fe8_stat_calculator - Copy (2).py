#!/usr/bin/env python3
"""
Fire Emblem 8 Unit Stat Calculator
Parses event files and calculates unit stats based on character, class, and equipment data.
"""

import os
import re
import csv
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class DefinitionResolver:
    """Resolves definitions from TableEntryDefinitions.event and FE8Definitions.txt"""
    
    def __init__(self):
        self.definitions: Dict[str, str] = {}
    
    def load_definitions(self, filepath: str):
        """Load definitions from a file"""
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                # Match #define lines
                match = re.match(r'#define\s+(\w+)\s+(.+)', line)
                if match:
                    name, value = match.groups()
                    self.definitions[name] = value.strip()
    
    def resolve(self, value: str) -> str:
        """Resolve a definition to its final value (byte or hex)"""
        if not value:
            return value
        
        # If it's already a hex value, return it
        if value.startswith('0x') or value.startswith('0X'):
            return value
        
        # If it's a number, return it
        if value.isdigit():
            return value
        
        # Otherwise, try to resolve it
        seen = set()
        current = value
        
        while current in self.definitions and current not in seen:
            seen.add(current)
            next_val = self.definitions[current]
            
            # If we found a hex value or number, return it
            if next_val.startswith('0x') or next_val.startswith('0X'):
                return next_val
            if next_val.isdigit():
                return next_val
            
            current = next_val
        
        # Return the final value even if not fully resolved
        return current if current in self.definitions else value
    
    def to_int(self, value: str) -> int:
        """Convert a resolved value to an integer"""
        resolved = self.resolve(value)
        if resolved.startswith('0x') or resolved.startswith('0X'):
            return int(resolved, 16)
        return int(resolved)


class UnitParser:
    """Parses UNIT entries from event files"""
    
    UNIT_PATTERN = re.compile(
        r'UNIT\s+(\w+)\s+(\w+)\s+(\w+)\s+Level\((\d+),(\w+),(\w+)\)\s+'
        r'\[[\d,]+\]\s+\w+\s+\w+\s+\w+\s+\w+\s+'
        r'\[([^\]]+)\]'
    )
    
    @staticmethod
    def parse_unit(line: str) -> Optional[Dict]:
        """Parse a UNIT line and extract relevant fields"""
        match = UnitParser.UNIT_PATTERN.search(line)
        if not match:
            return None
        
        char_id, class_id, _, level, allegiance, autolevel, items = match.groups()
        
        # Parse items
        item_list = [item.strip() for item in items.split(',')]
        # Filter out 0x0 entries
        item_list = [item for item in item_list if item != '0x0']
        
        return {
            'char_id': char_id,
            'class_id': class_id,
            'level': int(level),
            'allegiance': allegiance,
            'autolevel': autolevel.lower() == 'true',
            'items': item_list[:4]  # Max 4 items
        }
    
    @staticmethod
    def find_units_in_file(filepath: str) -> List[Dict]:
        """Find all UNIT entries in a file"""
        units = []
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                unit = UnitParser.parse_unit(line)
                if unit:
                    units.append(unit)
        return units


class TableLoader:
    """Loads and queries CSV tables"""
    
    def __init__(self):
        self.tables: Dict[str, List[Dict]] = {}
    
    def load_csv(self, filepath: str, table_name: str):
        """Load a CSV file into memory"""
        rows = []
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        self.tables[table_name] = rows
    
    def find_by_id(self, table_name: str, id_value: str, id_column: str = 'ID') -> Optional[Dict]:
        """Find a row in a table by ID"""
        if table_name not in self.tables:
            return None
        
        # Normalize hex values for comparison
        def normalize_id(val):
            if not val:
                return val
            val = val.strip()
            if val.startswith('0x') or val.startswith('0X'):
                try:
                    # Convert to int and back to normalized hex
                    return hex(int(val, 16))
                except:
                    return val
            return val
        
        normalized_search = normalize_id(id_value)
        
        for row in self.tables[table_name]:
            row_id = normalize_id(row.get(id_column, ''))
            if row_id == normalized_search:
                return row
        return None
    
    def find_by_row_number(self, table_name: str, row_number: int) -> Optional[Dict]:
        """Find a row in a table by row number (0-indexed, after header)"""
        if table_name not in self.tables:
            return None
        if row_number < 0 or row_number >= len(self.tables[table_name]):
            return None
        return self.tables[table_name][row_number]
    
    def find_character_by_hex_id(self, hex_id: str) -> Optional[Dict]:
        """Find a character by hex ID in the first column (0x803d64)"""
        if 'character' not in self.tables:
            return None
        
        # Normalize the hex ID
        normalized_search = hex_id.strip().lower()
        if not normalized_search.startswith('0x'):
            return None
        
        for row in self.tables['character']:
            first_col = row.get('0x803d64', '').strip()
            if first_col:
                # Extract hex ID from format "0xNN Name"
                parts = first_col.split()
                if parts and parts[0].lower() == normalized_search:
                    return row
        
        return None
    
    def get_stat_value(self, row: Dict, column: str, default: int = 0) -> int:
        """Get a stat value from a row, handling empty strings"""
        value = row.get(column, '').strip()
        if not value or value == '':
            return default
        try:
            return int(value)
        except ValueError:
            return default


class StatCalculator:
    """Calculates unit stats based on character, class, and equipment"""
    
    def __init__(self, resolver: DefinitionResolver, table_loader: TableLoader):
        self.resolver = resolver
        self.tables = table_loader
    
    @staticmethod
    def parse_difficulty_bonus(difficulty_hex: str) -> dict:
        """
        Parse difficulty bonus from hex format: 0x[Normal][Hard][Easy]
        
        Format interpretation:
        - Position 0 (Normal): Autolevels to SUBTRACT (e.g., 2 = -2 autolevels)
        - Position 1 (Hard): Autolevels to ADD (e.g., 3 = +3 autolevels)
        - Position 2 (Easy): Autolevels to SUBTRACT (e.g., 4 = -4 autolevels)
        
        Each digit can be 0-F (0-15).
        
        Example: 0x234 = Normal: -2, Hard: +3, Easy: -4
        Example: 0x050 = Normal: 0, Hard: +5, Easy: 0
        """
        if not difficulty_hex or not difficulty_hex.startswith('0x'):
            return {'Normal': 0, 'Hard': 0, 'Easy': 0}
        
        try:
            # Remove 0x prefix and ensure we have 3 digits
            hex_str = difficulty_hex[2:].strip().upper()
            if len(hex_str) < 3:
                hex_str = hex_str.zfill(3)
            
            # Parse each digit
            normal_hex = hex_str[0]
            hard_hex = hex_str[1]
            easy_hex = hex_str[2]
            
            # Normal and Easy are subtractions, Hard is addition
            normal_val = -int(normal_hex, 16)  # Negative (subtract)
            hard_val = int(hard_hex, 16)       # Positive (add)
            easy_val = -int(easy_hex, 16)      # Negative (subtract)
            
            return {
                'Normal': normal_val,
                'Hard': hard_val,
                'Easy': easy_val
            }
        except (ValueError, IndexError):
            return {'Normal': 0, 'Hard': 0, 'Easy': 0}
    
    def calculate_unit_stats(self, unit: Dict, chapter_name: str, difficulty_bonus: int = 0) -> Dict:
        """Calculate all stats for a unit
        
        Args:
            unit: Unit data dictionary
            chapter_name: Name of the chapter
            difficulty_bonus: Level adjustment from difficulty mode (added to chapLevel for calculations)
        """
        # For character lookup, try the string name first, then resolved hex ID
        char_name = unit['char_id']
        char_id_hex = self.resolver.resolve(unit['char_id'])
        
        # For class lookup, try string name first, then resolved hex ID
        class_name = unit['class_id']
        class_id_hex = self.resolver.resolve(unit['class_id'])
        
        # Get character and class data
        # For characters: try name first, then hex ID in ID column, then hex ID in first column
        char_data = self.tables.find_by_id('character', char_name)
        if not char_data:
            char_data = self.tables.find_by_id('character', char_id_hex)
        if not char_data:
            # Try searching by hex ID in the first column (for entries like "0x80 Soldier")
            char_data = self.tables.find_character_by_hex_id(char_id_hex)
        
        class_data = self.tables.find_by_id('class', class_name)
        if not class_data:
            # Try hex ID
            class_data = self.tables.find_by_id('class', class_id_hex)
        
        if not char_data or not class_data:
            return None
        
        # For mag tables, extract the hex ID from the first column of the main table
        # The first column format is "0xNN Name" - extract the hex part
        char_first_col = char_data.get('0x803d64', '').strip()
        char_hex_id = None
        if char_first_col:
            # Extract hex ID (format: "0x01 Hero" -> "0x01")
            parts = char_first_col.split()
            if parts and parts[0].startswith('0x'):
                try:
                    char_hex_id = int(parts[0], 16)
                except:
                    pass
        
        class_first_col = class_data.get('INLINE NewClassTable', '').strip()
        class_hex_id = None
        if class_first_col:
            # For classes, some have string IDs, some have hex
            # Try to parse as hex if it starts with 0x
            if class_first_col.startswith('0x'):
                try:
                    class_hex_id = int(class_first_col.split()[0], 16)
                except:
                    pass
            else:
                # String ID like "Veteran" - need to resolve it
                try:
                    resolved = self.resolver.resolve(class_name)
                    if resolved and resolved.startswith('0x'):
                        class_hex_id = int(resolved, 16)
                except:
                    pass
        
        # Use hex ID as row index into mag tables
        mag_char_data = None
        if char_hex_id is not None:
            mag_char_data = self.tables.find_by_row_number('mag_char', char_hex_id)
        
        mag_class_data = None
        if class_hex_id is not None:
            mag_class_data = self.tables.find_by_row_number('mag_class', class_hex_id)
        
        # Get base stats
        char_bases = {
            'HP': self.tables.get_stat_value(char_data, 'HP'),
            'Atk': self.tables.get_stat_value(char_data, 'Atk'),
            'Skl': self.tables.get_stat_value(char_data, 'Skl'),
            'Spd': self.tables.get_stat_value(char_data, 'Spd'),
            'Def': self.tables.get_stat_value(char_data, 'Def'),
            'Res': self.tables.get_stat_value(char_data, 'Res'),
            'Luck': self.tables.get_stat_value(char_data, 'Luck'),
            'Con': self.tables.get_stat_value(char_data, 'Con'),
        }
        
        class_bases = {
            'HP': self.tables.get_stat_value(class_data, 'HP'),
            'Atk': self.tables.get_stat_value(class_data, 'Atk'),
            'Skl': self.tables.get_stat_value(class_data, 'Skl'),
            'Spd': self.tables.get_stat_value(class_data, 'Spd'),
            'Def': self.tables.get_stat_value(class_data, 'Def'),
            'Res': self.tables.get_stat_value(class_data, 'Res'),
            'Con': self.tables.get_stat_value(class_data, 'Con'),
            'Luck': 0,  # Classes don't have Luck bases
        }
        
        # Get magic bases
        char_mag = self.tables.get_stat_value(mag_char_data, 'Base Magic') if mag_char_data else 0
        class_mag = self.tables.get_stat_value(mag_class_data, 'Base Magic') if mag_class_data else 0
        
        # Get growths
        class_growths = {
            'HP': self.tables.get_stat_value(class_data, 'HP Growth'),
            'Atk': self.tables.get_stat_value(class_data, 'Atk Growth'),
            'Skl': self.tables.get_stat_value(class_data, 'Skl Growth'),
            'Spd': self.tables.get_stat_value(class_data, 'Spd Growth'),
            'Def': self.tables.get_stat_value(class_data, 'Def Growth'),
            'Res': self.tables.get_stat_value(class_data, 'Res Growth'),
            'Luck': self.tables.get_stat_value(class_data, 'Luck Growth'),
        }
        
        class_mag_growth = self.tables.get_stat_value(mag_class_data, 'Magic Growth') if mag_class_data else 0
        
        # Check if promoted
        class_ability = class_data.get('CharClassAbility', '')
        is_promoted = 'IsPromoted' in class_ability
        
        # Calculate final stats
        level = unit['level']
        allegiance = unit['allegiance']
        autolevel = unit['autolevel']
        
        # Apply difficulty bonus to level for calculations (but keep original level for output)
        adjusted_level = level + difficulty_bonus
        
        stats = {}
        
        if allegiance in ['Enemy', 'NPC']:
            # Enemy/NPC units use only class growths
            for stat in ['HP', 'Atk', 'Skl', 'Spd', 'Def', 'Res', 'Luck']:
                base = class_bases[stat] + char_bases[stat]
                if autolevel:
                    growth_bonus = class_growths[stat] * adjusted_level // 100
                    if is_promoted:
                        growth_bonus += class_growths[stat] * 10 // 100
                    base += growth_bonus
                stats[stat] = base
            
            # Magic
            mag_base = class_mag + char_mag
            if autolevel:
                mag_growth_bonus = class_mag_growth * adjusted_level // 100
                if is_promoted:
                    mag_growth_bonus += class_mag_growth * 10 // 100
                mag_base += mag_growth_bonus
            stats['Mag'] = mag_base
            
        else:  # Ally
            # Ally units use character growths
            char_growths = {
                'HP': self.tables.get_stat_value(char_data, 'HP Growth'),
                'Atk': self.tables.get_stat_value(char_data, 'Atk Growth'),
                'Skl': self.tables.get_stat_value(char_data, 'Skl Growth'),
                'Spd': self.tables.get_stat_value(char_data, 'Spd Growth'),
                'Def': self.tables.get_stat_value(char_data, 'Def Growth'),
                'Res': self.tables.get_stat_value(char_data, 'Res Growth'),
                'Luck': self.tables.get_stat_value(char_data, 'Luck Growth'),
            }
            
            char_mag_growth = self.tables.get_stat_value(mag_char_data, 'Magic Growth') if mag_char_data else 0
            
            for stat in ['HP', 'Atk', 'Skl', 'Spd', 'Def', 'Res', 'Luck']:
                base = class_bases[stat] + char_bases[stat]
                if autolevel:
                    growth_bonus = char_growths[stat] * adjusted_level // 100
                    base += growth_bonus
                stats[stat] = base
            
            # Magic
            mag_base = class_mag + char_mag
            if autolevel:
                mag_growth_bonus = char_mag_growth * adjusted_level // 100
                mag_base += mag_growth_bonus
            stats['Mag'] = mag_base
        
        # Con is always base only
        stats['Con'] = class_bases['Con'] + char_bases['Con']
        stats['Lv'] = level
        
        # Prepare result
        result = {
            'UnitID': f"{unit['char_id']}_{unit['class_id']}",
            'Chapter': chapter_name,
            'Allegiance': allegiance,  # Add allegiance for color coding
            'Lv': stats['Lv'],
            'HP': stats['HP'],
            'Str': stats['Atk'],  # 'Atk' in game tables = 'Str' in output format
            'Mag': stats['Mag'],
            'Skl': stats['Skl'],
            'Spd': stats['Spd'],
            'Def': stats['Def'],
            'Res': stats['Res'],
            'Luck': stats['Luck'],
            'Con': stats['Con'],
            'weapons': []
        }
        
        # Calculate weapon stats
        for item in unit['items']:
            weapon_stats = self.calculate_weapon_stats(item, result)
            if weapon_stats:
                result['weapons'].append(weapon_stats)
        
        return result
    
    def calculate_weapon_stats(self, item_id: str, unit_result: Dict) -> Optional[Dict]:
        """Calculate stats for a weapon"""
        # Resolve item ID
        resolved_id = self.resolver.resolve(item_id)
        
        # Get item data
        item_data = self.tables.find_by_id('item', resolved_id)
        if not item_data:
            return None
        
        # Check if it's a weapon
        item_ability = item_data.get('ItemAbility', '')
        if 'IsWeapon' not in item_ability:
            return None
        
        # Get weapon stats
        mt = self.tables.get_stat_value(item_data, 'Mt')
        hit = self.tables.get_stat_value(item_data, 'Hit')
        wt = self.tables.get_stat_value(item_data, 'Wt')
        crit = self.tables.get_stat_value(item_data, 'Crit')
        
        # Calculate derived stats
        is_magic = 'MagicDamage' in item_ability or 'IsMagic' in item_ability
        atk = mt + (unit_result['Mag'] if is_magic else unit_result['Str'])
        
        weapon_hit = 2 * unit_result['Skl'] + unit_result['Luck'] + hit
        weapon_crit = crit
        
        # AS = Spd - max(0, Wt - Con)
        wt_penalty = max(0, wt - unit_result['Con'])
        attack_speed = unit_result['Spd'] - wt_penalty
        
        avoid = 2 * unit_result['Spd'] + unit_result['Luck']
        
        return {
            'Weapon': item_id,
            'Atk': atk,
            'Hit': weapon_hit,
            'Crit': weapon_crit,
            'AS': attack_speed,
            'Avoid': avoid
        }


def export_to_excel(results: Dict[str, List[Dict]], output_path: str):
    """Export calculated stats to Excel with one sheet per chapter
    
    Features:
    - Units separated by allegiance (Ally, NPC, Enemy)
    - Color-coded rows (light blue for Ally, light green for NPC, light red for Enemy)
    - Additional columns: AS+4, Mag OHKO, Phys OHKO, Mag ORKO, Phys ORKO
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Allegiance colors
    ally_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
    npc_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")   # Light green
    enemy_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid") # Light red
    
    for chapter_name in sorted(results.keys()):
        units = results[chapter_name]
        if not units:
            continue
        
        # Separate units by allegiance
        ally_units = [u for u in units if u.get('Allegiance') == 'Ally']
        npc_units = [u for u in units if u.get('Allegiance') == 'NPC']
        enemy_units = [u for u in units if u.get('Allegiance') == 'Enemy']
        
        # Create sheet
        ws = wb.create_sheet(title=chapter_name[:31])  # Excel sheet name limit
        
        # Write header
        headers = ['UnitID', 'Lv', 'HP', 'Str', 'Mag', 'Skl', 'Spd', 'Def', 'Res', 'Luck', 'Con', 
                   'Weapon', 'Atk', 'Hit', 'Crit', 'AS', 'Avoid', 'AS+4', 
                   'Mag OHKO', 'Phys OHKO', 'Bomb OHKO', 'Mag ORKO', 'Phys ORKO']
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row_idx = 2
        
        # Helper function to write unit rows with color coding
        def write_unit_section(unit_list, fill_color):
            nonlocal row_idx
            for unit in unit_list:
                # Use actual HP (not HP Cap) for OHKO/ORKO calculations
                hp = unit['HP']
                def_stat = unit['Def']
                res_stat = unit['Res']
                
                if not unit['weapons']:
                    # Unit with no weapons
                    ws.cell(row=row_idx, column=1, value=unit['UnitID'])
                    ws.cell(row=row_idx, column=2, value=unit['Lv'])
                    ws.cell(row=row_idx, column=3, value=hp)
                    ws.cell(row=row_idx, column=4, value=unit['Str'])
                    ws.cell(row=row_idx, column=5, value=unit['Mag'])
                    ws.cell(row=row_idx, column=6, value=unit['Skl'])
                    ws.cell(row=row_idx, column=7, value=unit['Spd'])
                    ws.cell(row=row_idx, column=8, value=def_stat)
                    ws.cell(row=row_idx, column=9, value=res_stat)
                    ws.cell(row=row_idx, column=10, value=unit['Luck'])
                    ws.cell(row=row_idx, column=11, value=unit['Con'])
                    # Columns 12-17 are weapon stats (empty for units without weapons)
                    # Column 18: AS+4 (no weapon, so no AS)
                    # Columns 19-23: OHKO/ORKO stats
                    ws.cell(row=row_idx, column=19, value=hp + res_stat)  # Mag OHKO
                    ws.cell(row=row_idx, column=20, value=hp + def_stat)  # Phys OHKO
                    ws.cell(row=row_idx, column=21, value=int(hp * 0.75) + res_stat)  # Bomb OHKO
                    ws.cell(row=row_idx, column=22, value=hp // 2 + res_stat)  # Mag ORKO
                    ws.cell(row=row_idx, column=23, value=hp // 2 + def_stat)  # Phys ORKO
                    
                    # Apply color to entire row
                    for col in range(1, 24):
                        ws.cell(row=row_idx, column=col).fill = fill_color
                    
                    row_idx += 1
                else:
                    # Unit with weapons
                    for weapon in unit['weapons']:
                        ws.cell(row=row_idx, column=1, value=unit['UnitID'])
                        ws.cell(row=row_idx, column=2, value=unit['Lv'])
                        ws.cell(row=row_idx, column=3, value=hp)
                        ws.cell(row=row_idx, column=4, value=unit['Str'])
                        ws.cell(row=row_idx, column=5, value=unit['Mag'])
                        ws.cell(row=row_idx, column=6, value=unit['Skl'])
                        ws.cell(row=row_idx, column=7, value=unit['Spd'])
                        ws.cell(row=row_idx, column=8, value=def_stat)
                        ws.cell(row=row_idx, column=9, value=res_stat)
                        ws.cell(row=row_idx, column=10, value=unit['Luck'])
                        ws.cell(row=row_idx, column=11, value=unit['Con'])
                        ws.cell(row=row_idx, column=12, value=weapon['Weapon'])
                        ws.cell(row=row_idx, column=13, value=weapon['Atk'])
                        ws.cell(row=row_idx, column=14, value=weapon['Hit'])
                        ws.cell(row=row_idx, column=15, value=weapon['Crit'])
                        ws.cell(row=row_idx, column=16, value=weapon['AS'])
                        ws.cell(row=row_idx, column=17, value=weapon['Avoid'])
                        ws.cell(row=row_idx, column=18, value=weapon['AS'] + 4)  # AS+4
                        ws.cell(row=row_idx, column=19, value=hp + res_stat)  # Mag OHKO
                        ws.cell(row=row_idx, column=20, value=hp + def_stat)  # Phys OHKO
                        ws.cell(row=row_idx, column=21, value=int(hp * 0.75) + res_stat)  # Bomb OHKO
                        ws.cell(row=row_idx, column=22, value=hp // 2 + res_stat)  # Mag ORKO
                        ws.cell(row=row_idx, column=23, value=hp // 2 + def_stat)  # Phys ORKO
                        
                        # Apply color to entire row
                        for col in range(1, 24):
                            ws.cell(row=row_idx, column=col).fill = fill_color
                        
                        row_idx += 1
        
        # Write units by allegiance with spacing
        if ally_units:
            write_unit_section(ally_units, ally_fill)
            row_idx += 1  # Add blank row separator
        
        if npc_units:
            write_unit_section(npc_units, npc_fill)
            row_idx += 1  # Add blank row separator
        
        if enemy_units:
            write_unit_section(enemy_units, enemy_fill)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)

class Ch0Parser:
    """Parses Ch0.event to extract unit group definitions"""
    
    @staticmethod
    def parse_ch0_unit_groups(filepath: str) -> Dict[str, List[str]]:
        """
        Parse Ch0.event and extract unit groups (Ch1Units, Ch2Units, etc.)
        
        Returns:
            Dictionary mapping group names to lists of character IDs
            Example: {'Ch1Units': ['Hero', 'Teacher'], 'Ch2Units': ['Pillow']}
        """
        unit_groups = {}
        current_group = None
        
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                
                # Check for group label (e.g., "Ch1Units:")
                # Pattern: word starting with Ch followed by digit(s), ending with Units:
                group_match = re.match(r'^(Units+Ch\d):', line)
                if group_match:
                    current_group = group_match.group(1)
                    unit_groups[current_group] = []
                    continue
                
                # If we're in a group, look for UNIT declarations
                if current_group:
                    unit = UnitParser.parse_unit(line)
                    if unit:
                        char_id = unit['char_id']
                        # Add to current group if not already present
                        if char_id not in unit_groups[current_group]:
                            unit_groups[current_group].append(char_id)
        
        return unit_groups
    
    @staticmethod
    def build_party_composition(unit_groups: Dict[str, List[str]]) -> Dict[str, List[str]]:
        """
        Build cumulative party composition for each chapter.
        
        Args:
            unit_groups: Dictionary of group names to character IDs
            
        Returns:
            Dictionary mapping chapter names to cumulative character lists
            Example: {'Ch1': ['Hero', 'Teacher'], 'Ch2': ['Hero', 'Teacher', 'Pillow']}
        """
        party_composition = {}
        
        # Extract chapter numbers from group names and sort
        chapter_nums = []
        for group_name in unit_groups.keys():
            match = re.match(r'Ch(\d+)Units', group_name)
            if match:
                chapter_nums.append(int(match.group(1)))
        
        chapter_nums = sorted(set(chapter_nums))
        
        # Build cumulative lists
        cumulative_units = []
        for chap_num in chapter_nums:
            group_name = f'Ch{chap_num}Units'
            if group_name in unit_groups:
                # Add units from this chapter's group
                for char_id in unit_groups[group_name]:
                    if char_id not in cumulative_units:
                        cumulative_units.append(char_id)
            
            # Store cumulative list for this chapter
            party_composition[f'Ch{chap_num}'] = cumulative_units.copy()
        
        return party_composition

def main():
    """Main function"""
    import sys
    import os
    
    # Get base path from command line or use current directory
    if len(sys.argv) > 1:
        base_path = sys.argv[1]
    else:
        # Use current directory as base
        base_path = os.getcwd()
    
    # File paths with new defaults
    table_defs_path = os.path.join(base_path, 'Definitions', 'TableEntryDefinitions.event')
    fe8_defs_path = os.path.join(base_path, 'EventAssembler', 'EAStandardLibrary', 'FE8 Definitions.txt')
    chapter_events_dir = os.path.join(base_path, 'Events', 'ChapterEvents')
    char_class_dir = os.path.join(base_path, 'Tables', 'NightmareModules', 'CharactersClasses')
    items_dir = os.path.join(base_path, 'Tables', 'NightmareModules', 'Items')
    chapter_data_dir = os.path.join(base_path, 'Tables', 'NightmareModules', 'ChapterData')
    
    char_table_path = os.path.join(char_class_dir, 'CharacterTable.csv')
    class_table_path = os.path.join(char_class_dir, 'ClassTable.csv')
    mag_char_path = os.path.join(char_class_dir, 'MagCharEditor.csv')
    mag_class_path = os.path.join(char_class_dir, 'MagClassEditor.csv')
    item_table_path = os.path.join(items_dir, 'ItemTable.csv')
    chapter_data_path = os.path.join(chapter_data_dir, 'ChapterData.csv')
    
    # Check if files exist, if not try flat structure fallback
    if not os.path.exists(table_defs_path):
        print("Standard structure not found, trying flat structure...")
        table_defs_path = os.path.join(base_path, 'TableEntryDefinitions.event')
        fe8_defs_path = os.path.join(base_path, 'FE8 Definitions.txt')
        chapter_events_dir = base_path  # Look in base path
        char_table_path = os.path.join(base_path, 'CharacterTable.csv')
        class_table_path = os.path.join(base_path, 'ClassTable.csv')
        mag_char_path = os.path.join(base_path, 'MagCharEditor.csv')
        mag_class_path = os.path.join(base_path, 'MagClassEditor.csv')
        item_table_path = os.path.join(base_path, 'ItemTable.csv')
        chapter_data_path = os.path.join(base_path, 'ChapterData.csv')
    
    # Check if files exist
    required_files = [
        table_defs_path, fe8_defs_path, char_table_path, class_table_path,
        mag_char_path, mag_class_path, item_table_path
    ]
    
    # ChapterData is optional
    has_chapter_data = os.path.exists(chapter_data_path)
    if not has_chapter_data:
        print("Warning: ChapterData.csv not found - difficulty modes will not be available")
    
    missing_files = [f for f in required_files if not os.path.exists(f)]
    if missing_files:
        print("ERROR: Required files not found:")
        for f in missing_files:
            print(f"  - {f}")
        print(f"\nBase path: {base_path}")
        print("\nExpected structure (standard):")
        print("  \\Definitions\\TableEntryDefinitions.event")
        print("  \\EventAssembler\\EAStandardLibrary\\FE8 Definitions.txt")
        print("  \\Events\\ChapterEvents\\")
        print("  \\Tables\\NightmareModules\\CharactersClasses\\")
        print("  \\Tables\\NightmareModules\\Items\\")
        print("  \\Tables\\NightmareModules\\ChapterData\\ (optional)")
        print("\nOr flat structure:")
        print("  \\TableEntryDefinitions.event")
        print("  \\FE8 Definitions.txt")
        print("  \\CharacterTable.csv, ClassTable.csv, etc.")
        print("  \\ChapterData.csv (optional)")
        return
    
    # Initialize components
    print("Loading definitions...")
    resolver = DefinitionResolver()
    resolver.load_definitions(table_defs_path)
    resolver.load_definitions(fe8_defs_path)
    
    print("Loading tables...")
    table_loader = TableLoader()
    table_loader.load_csv(char_table_path, 'character')
    table_loader.load_csv(class_table_path, 'class')
    table_loader.load_csv(mag_char_path, 'mag_char')
    table_loader.load_csv(mag_class_path, 'mag_class')
    table_loader.load_csv(item_table_path, 'item')
    
    # Load chapter data if available
    chapter_data_map = {}
    if has_chapter_data:
        table_loader.load_csv(chapter_data_path, 'chapter_data')
        # Build a map of EventsID -> chapter data
        for row in table_loader.tables['chapter_data']:
            events_id = row.get('EventsID', '').strip()
            if events_id:
                chapter_data_map[events_id] = row
    
    calculator = StatCalculator(resolver, table_loader)
    
    # Parse Ch0.event for unit groups
    ch0_path = os.path.join(chapter_events_dir, 'Ch0.event')
    party_composition = {}
    
    # Parse Ch0.event for unit groups
    ch0_path = os.path.join(chapter_events_dir, 'Ch0.event')
    party_composition = {}
    
    if os.path.exists(ch0_path):
        print("Parsing Ch0.event for unit groups...")
        unit_groups = Ch0Parser.parse_ch0_unit_groups(ch0_path)
        
        # Debug output
        print(f"  Found {len(unit_groups)} unit groups")
        if not unit_groups:
            print("  WARNING: No unit groups found in Ch0.event!")
            print("  Expected format: 'Ch1Units:', 'Ch2Units:', etc.")
        
        party_composition = Ch0Parser.build_party_composition(unit_groups)
        
        print("\nUnit Groups found:")
        for group_name in sorted(unit_groups.keys(), key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0):
            chars = unit_groups[group_name]
            print(f"  {group_name}: {', '.join(chars)}")
        
        print("\nParty Composition by Chapter:")
        for chap_name in sorted(party_composition.keys(), key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 0):
            chars = party_composition[chap_name]
            print(f"  {chap_name}: {', '.join(chars)}")
        print()
    else:
        print(f"Warning: Ch0.event not found at {ch0_path}")
        print("Party composition tracking will be disabled.\n")
    
    # Determine which difficulty modes to process
    include_easy = False  # Default: off
    # Check for command line argument
    if len(sys.argv) > 2 and sys.argv[2].lower() in ['--easy', '-e', 'easy']:
        include_easy = True
    
    difficulty_modes = ['Normal', 'Hard']
    if include_easy:
        difficulty_modes.append('Easy')
    
    # Parse units from all chapter event files
    print("Parsing chapter events...")
    results = {}
    
    if os.path.isdir(chapter_events_dir):
        # Process all .event files in the directory
        event_files = [f for f in os.listdir(chapter_events_dir) if f.endswith('.event')]
        
        for filename in sorted(event_files):
            chapter_file = os.path.join(chapter_events_dir, filename)
            chapter_name = filename[:-6]  # Remove .event extension
            
            # Skip Ch0
            if chapter_name == 'Ch0':
                continue
            
            units = UnitParser.find_units_in_file(chapter_file)
            
            # Remove duplicates
            unique_units = {}
            for unit in units:
                key = f"{unit['char_id']}_{unit['class_id']}"
                if key not in unique_units:
                    unique_units[key] = unit
            
            units = list(unique_units.values())
            
            if units:
                print(f"  {chapter_name}: {len(units)} unique units")
                
                # Show party composition for this chapter
                if chapter_name in party_composition:
                    party_units = party_composition[chapter_name]
                    print(f"    Party at {chapter_name}: {', '.join(party_units)}")
                else:
                    print(f"    Party at {chapter_name}: (no party data)")
                
                # Get chapter data for difficulty bonuses
                events_id = f"{chapter_name}EventsID"
                chapter_data = chapter_data_map.get(events_id, {})
                difficulty_hex = chapter_data.get('DifficultyBonus', '0x0')
                difficulty_bonuses = calculator.parse_difficulty_bonus(difficulty_hex)
                
                # Calculate stats for each difficulty mode
                for difficulty in difficulty_modes:
                    sheet_name = f"{chapter_name} {difficulty}"
                    difficulty_bonus = difficulty_bonuses.get(difficulty, 0)
                    
                    results[sheet_name] = []
                    
                    for unit in units:
                        stats = calculator.calculate_unit_stats(unit, chapter_name, difficulty_bonus)
                        if stats:
                            results[sheet_name].append(stats)
    
    if not results:
        print("ERROR: No chapter event files found or no units parsed!")
        return
    
    # Export to Excel in the same directory as the script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, 'unit_stats.xlsx')
    
    print(f"Calculating stats...")
    print(f"Exporting to {output_path}...")
    export_to_excel(results, output_path)
    print("Done!")
    print(f"\nOutput file: {output_path}")


if __name__ == '__main__':
    main()
